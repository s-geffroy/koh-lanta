#!/usr/bin/env python3
"""Compare l'origine geographique des aventuriers a la population francaise.

Meme methode que `prenoms_insee.py`, et meme prudence : un effectif brut ne dit
rien tant qu'on ne l'a pas rapporte a ce qu'on attendrait. Vingt-quatre
aventuriers parisiens, est-ce beaucoup ? Seule la population de Paris repond.

La reference est le tableau DEP2 de l'INSEE -- population au 1er janvier par
departement, sexe et groupe d'ages, series depuis 1990 -- sous licence ouverte.
On y prend la tranche **20-59 ans**, celle qui recouvre le mieux les
aventuriers (18-66 ans, mediane 33), et l'annee de CHAQUE saison : la France de
2003 n'est pas celle de 2026.

    tools/atelier python3 tools/extraction/geographie_insee.py --ecrire
"""
import argparse
import collections
import io
import os
import sys
import urllib.request

import numpy as np
import openpyxl
import yaml

import modeles   # pour fabriquer les tests dans la forme du registre commun

RACINE = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", ".."))
SOURCE = ("https://www.insee.fr/fr/statistiques/fichier/8560704/"
          "2_Pop1janv_grages_depreg.xlsx")
EXTRAIT = os.path.join(RACINE, "specs", "sources", "insee", "population-dep-extrait.csv")
SORTIE = os.path.join(RACINE, "_data", "geographie.yml")

COLONNE_20_59 = 5      # « De 20 a 59 ans », ensemble hommes et femmes
DERNIERE_ANNEE = 2025  # au-dela, le fichier ne va pas : on reconduit la derniere

# Les deux departements corses sont tantot distingues, tantot fondus dans nos
# sources. On les fond, systematiquement, des deux cotes.
CORSE = {"Corse-du-Sud", "Haute-Corse", "Corse"}

TIRAGES = 10_000
GRAINE = 20260828


def telecharger():
    return urllib.request.urlopen(SOURCE, timeout=120).read()


# Ce qui, dans la colonne « Code REG-DEP », n'est pas un departement : les
# regions (deux caracteres), et les trois agregats nationaux.
AGREGATS = {"FR", "FM", "97"}


def lire_populations(donnees, annees):
    """Population 20-59 par departement et par region, annee par annee.

    Le code d'un departement metropolitain vaut REGION + DEPARTEMENT, sur
    quatre caracteres (Paris : 1175, region 11). Les departements d'outre-mer,
    eux, n'ont pas de prefixe de region : leur code tient sur trois chiffres
    (Martinique : 972). Le filtre doit accepter les deux, sans quoi les DOM
    disparaissent -- ce qui etait le cas de la premiere version.
    """
    classeur = openpyxl.load_workbook(io.BytesIO(donnees), read_only=True, data_only=True)
    par_annee, regions_par_annee, totaux, region_de = {}, {}, {}, {}
    noms_region = {}
    for annee in sorted(annees):
        feuille = str(min(annee, DERNIERE_ANNEE))
        if feuille not in classeur.sheetnames:
            continue
        deps, regions, total = {}, {}, None
        for ligne in classeur[feuille].iter_rows(values_only=True):
            nom, code = ligne[0], ligne[1]
            if not nom or not code:
                continue
            nom, code = str(nom).strip(), str(code).strip()
            valeur = ligne[COLONNE_20_59]
            if valeur is None:
                continue
            if code == "FR":
                total = int(valeur)
            elif code in AGREGATS:
                continue
            elif len(code) == 2:
                regions[nom] = int(valeur)
                noms_region[code] = nom
            elif len(code) >= 3:
                cle = "Corse" if nom in CORSE else nom
                deps[cle] = deps.get(cle, 0) + int(valeur)
                # Les annees sont parcourues dans l'ordre : la DERNIERE vue
                # l'emporte, ce qui range chaque departement dans sa region
                # actuelle et non dans celle d'avant la reforme de 2016.
                prefixe = code[:2] if len(code) >= 4 else code
                region_de[cle] = prefixe
        if total:
            par_annee[annee] = deps
            regions_par_annee[annee] = regions
            totaux[annee] = total
    # Les DOM n'ont pas de ligne de region : chacun est sa propre region.
    for dep, prefixe in list(region_de.items()):
        if prefixe not in noms_region:
            noms_region[prefixe] = dep
    appartenance = {d: noms_region.get(pref, d) for d, pref in region_de.items()}
    for annee, regions in regions_par_annee.items():
        for dep, pop in par_annee[annee].items():
            r = appartenance.get(dep)
            if r and r not in regions:
                regions[r] = regions.get(r, 0) + pop
    return par_annee, totaux, regions_par_annee, appartenance


def main():
    args = argparse.ArgumentParser(description=__doc__)
    args.add_argument("--ecrire", action="store_true")
    args = args.parse_args()

    lire = lambda n: yaml.safe_load(open(os.path.join(RACINE, "_data", n), encoding="utf-8"))
    saisons = {s["id"]: s for s in lire("saisons.yml")}
    parts = lire("participations.yml")

    # Les participations localisees, hors saisons speciales : les revenants y
    # comptent deux fois et fausseraient la geographie du recrutement.
    retenues = []
    for p in parts:
        s = saisons.get(p["saison"]) or {}
        if s.get("annulee") or s.get("speciale") or not p.get("localisation"):
            continue
        retenues.append((p["localisation"].strip(), s.get("annee")))

    annees = sorted({a for _, a in retenues if a})
    print(f"participations localisees (saisons classiques) : {len(retenues)}")
    print(f"annees concernees : {min(annees)}-{max(annees)}")

    donnees = telecharger()
    print(f"telecharge : {len(donnees)} octets depuis l'INSEE")
    populations, totaux, regions, appartenance = lire_populations(donnees, annees)
    print(f"annees lues dans le tableau INSEE : {len(populations)}")

    connus = set()
    for deps in populations.values():
        connus |= set(deps)

    observe = collections.Counter()
    hors_france, inconnus = collections.Counter(), collections.Counter()
    utilises = []
    for lieu, annee in retenues:
        cle = "Corse" if lieu in CORSE else lieu
        if cle not in connus:
            # Un lieu absent du fichier INSEE est soit etranger, soit mal
            # orthographie. On ne devine pas : on le compte a part et on le dit.
            (hors_france if annee else inconnus)[lieu] += 1
            inconnus[lieu] += 0
            hors_france[lieu] += 0
            continue
        observe[cle] += 1
        utilises.append((cle, annee))

    ecartes = [l for l, a in retenues if ("Corse" if l in CORSE else l) not in connus]
    print(f"lieux hors du fichier INSEE : {len(ecartes)} participations "
          f"({', '.join(sorted(set(ecartes)))})")
    print(f"participations comparables : {len(utilises)}")

    # L'attendu : pour chaque aventurier, la probabilite qu'une personne de
    # 20-59 ans tiree au hasard en France l'annee de sa saison habite ce
    # departement. On somme sur tout le casting, jamais sur les seuls porteurs.
    attendu = collections.defaultdict(float)
    poids = {}
    for annee in sorted({a for _, a in utilises}):
        n = sum(1 for _, a in utilises if a == annee)
        deps, total = populations[annee], totaux[annee]
        poids[annee] = (deps, total, n)
        for d, pop in sorted(deps.items()):
            attendu[d] += n * pop / total

    # Le test : on tire len(utilises) personnes dans la France de leur annee,
    # dix mille fois, et on compare la dispersion obtenue a celle qu'on observe.
    rng = np.random.default_rng(np.random.SeedSequence([GRAINE, 7]))
    noms = sorted(set(attendu) | set(observe))
    index = {d: i for i, d in enumerate(noms)}
    vecteur_attendu = np.array([attendu[d] for d in noms])
    vecteur_observe = np.zeros(len(noms))
    for d, n in observe.items():
        vecteur_observe[index[d]] = n

    def dispersion(v):
        with np.errstate(divide="ignore", invalid="ignore"):
            x = (v - vecteur_attendu) ** 2 / np.where(vecteur_attendu > 0,
                                                      vecteur_attendu, np.nan)
        return float(np.nansum(x))

    observe_stat = dispersion(vecteur_observe)
    nulle = np.zeros(TIRAGES)
    probas = {}
    for annee, (deps, total, n) in poids.items():
        p = np.zeros(len(noms))
        for d, pop in deps.items():
            if d in index:
                p[index[d]] = pop / total
        probas[annee] = (p / p.sum(), n)
    for t in range(TIRAGES):
        v = np.zeros(len(noms))
        for p, n in probas.values():
            v += rng.multinomial(n, p)
        nulle[t] = dispersion(v)
    test_dep = modeles._test(
        "geographie_departements", "L'origine géographique, par département",
        "Les aventuriers viennent-ils de départements plus variés, ou plus "
        "concentrés, qu'un tirage dans la population française ne le donnerait ?",
        observe_stat, nulle, unite="",
        lecture="Dispersion entre les effectifs observés et ceux qu'on attendrait "
                "d'un tirage au hasard dans la population de 20 à 59 ans, année par "
                "année. Au-dessus de l'attendu : le recrutement n'épouse pas la "
                "démographie.")
    p_value = test_dep["p"]

    lignes = []
    for d in noms:
        a = attendu[d]
        lignes.append({
            "departement": d,
            "observe": int(vecteur_observe[index[d]]),
            "attendu": round(float(a), 2),
            "indice": round(float(vecteur_observe[index[d]]) / float(a), 2)
                      if a >= 1 else None,
        })
    lignes.sort(key=lambda x: (-(x["indice"] or -1), -x["observe"], x["departement"]))

    print()
    print(f"dispersion observee : {observe_stat:.1f}  |  attendue au hasard : "
          f"{nulle.mean():.1f}  |  p = {p_value:.4f}")
    print("les plus sur-representes :")
    for x in [y for y in lignes if y["indice"]][:6]:
        print(f"   {x['departement']:24s} {x['observe']:>3} pour {x['attendu']:>6} "
              f"attendus  x{x['indice']}")
    print("les plus sous-representes :")
    for x in [y for y in lignes if y["indice"]][-6:]:
        print(f"   {x['departement']:24s} {x['observe']:>3} pour {x['attendu']:>6} "
              f"attendus  x{x['indice']}")

    # --- le niveau regional ------------------------------------------------
    # Quatre-vingt-dix departements pour trois cents personnes, c'est trois
    # personnes par departement : un indice y est un bruit, pas un resultat. La
    # region, elle, en compte une vingtaine -- assez pour dire quelque chose.
    from scipy import stats as st

    obs_reg = collections.Counter()
    for d, a in utilises:
        obs_reg[appartenance.get(d, d)] += 1
    att_reg = collections.defaultdict(float)
    for annee, (deps, total, n) in poids.items():
        for d, pop in sorted(deps.items()):
            att_reg[appartenance.get(d, d)] += n * pop / total

    n_total = len(utilises)
    regions_lignes = []
    for r in sorted(set(att_reg) | set(obs_reg)):
        a = att_reg[r]
        o = obs_reg[r]
        if a < 1:
            continue
        bas, haut = st.binomtest(o, n_total).proportion_ci(confidence_level=0.95)
        # PyYAML ne sait pas ecrire un np.float64 : on repasse en flottant
        # Python avant de verser quoi que ce soit dans _data/.
        regions_lignes.append({
            "region": r, "observe": int(o), "attendu": round(float(a), 1),
            "indice": round(float(o) / float(a), 2),
            "indice_bas": round(float(bas) * n_total / float(a), 2),
            "indice_haut": round(float(haut) * n_total / float(a), 2),
        })
    regions_lignes.sort(key=lambda x: -x["indice"])

    vec_o = np.array([x["observe"] for x in regions_lignes], dtype=float)
    vec_a = np.array([x["attendu"] for x in regions_lignes], dtype=float)
    stat_reg = float(np.sum((vec_o - vec_a) ** 2 / vec_a))
    rng2 = np.random.default_rng(np.random.SeedSequence([GRAINE, 11]))
    ordre_reg = {x["region"]: i for i, x in enumerate(regions_lignes)}
    nulle_reg = np.zeros(TIRAGES)
    probas_reg = {}
    for annee, (deps, total, n) in poids.items():
        p_ = np.zeros(len(regions_lignes))
        for d, pop in deps.items():
            r = appartenance.get(d, d)
            if r in ordre_reg:
                p_[ordre_reg[r]] += pop / total
        probas_reg[annee] = (p_ / p_.sum(), n)
    for t in range(TIRAGES):
        v = np.zeros(len(regions_lignes))
        for p_, n in probas_reg.values():
            v += rng2.multinomial(n, p_)
        nulle_reg[t] = float(np.sum((v - vec_a) ** 2 / vec_a))
    test_reg = modeles._test(
        "geographie_regions", "L'origine géographique, par région",
        "À l'échelle des régions — assez peuplées pour trancher — le recrutement "
        "suit-il la démographie française ?",
        stat_reg, nulle_reg, unite="",
        lecture="Même mesure, agrégée par région. C'est le niveau où la question "
                "se décide : un département compte trois aventuriers, une région "
                "une vingtaine.")
    p_reg = test_reg["p"]

    print()
    print(f"REGIONS — dispersion observee {stat_reg:.1f} contre {nulle_reg.mean():.1f} "
          f"attendue, p = {p_reg:.4f}")
    for x in regions_lignes:
        print(f"   {x['region']:28s} {x['observe']:>3} pour {x['attendu']:>5} "
              f"attendus  x{x['indice']:<5} [{x['indice_bas']} ; {x['indice_haut']}]")

    sortie = {
        "source": SOURCE,
        "champ": "population de 20 a 59 ans, au 1er janvier de l'annee de chaque saison",
        "participations": len(utilises),
        "hors_fichier": sorted(set(ecartes)),
        "nb_hors_fichier": len(ecartes),
        "annees": [min(annees), max(annees)],
        "derniere_annee_insee": DERNIERE_ANNEE,
        "departements_classes": sum(1 for x in lignes if x["indice"]),
        "dispersion_observee": round(float(observe_stat), 1),
        "dispersion_attendue": round(float(nulle.mean()), 1),
        "p": round(float(p_value), 4),
        "tirages": TIRAGES,
        "departements": lignes,
        "regions": regions_lignes,
        "regions_dispersion_observee": round(float(stat_reg), 1),
        "regions_dispersion_attendue": round(float(nulle_reg.mean()), 1),
        "regions_p": round(float(p_reg), 4),
        "tests": [test_dep, test_reg],
    }

    if args.ecrire:
        os.makedirs(os.path.dirname(EXTRAIT), exist_ok=True)
        with open(EXTRAIT, "w", encoding="utf-8") as f:
            f.write("annee;departement;population_20_59\n")
            for annee in sorted(populations):
                for d, pop in sorted(populations[annee].items()):
                    f.write(f"{annee};{d};{pop}\n")
        print(f"\necrit : {EXTRAIT}")
        with open(SORTIE, "w", encoding="utf-8") as f:
            f.write("# Fichier genere par tools/extraction/geographie_insee.py.\n"
                    "# Ne pas editer a la main : toute modification sera ecrasee.\n")
            yaml.safe_dump(sortie, f, allow_unicode=True, sort_keys=False, width=100)
        print(f"ecrit : {SORTIE}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
